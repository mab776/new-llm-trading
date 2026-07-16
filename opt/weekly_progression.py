"""Week-by-week equity progression export to Excel.

Replays the shared-portfolio simulator (same engine that produces the headline
842,919x / 72.9Tx multiples) but records the equity curve resampled to weekly
buckets, so the compounding path — not just the final multiple — is visible.

Six studies: {standard, aggressive} x {BTC+ETH+SOL (default), BTC only,
ETH only}. Single-asset studies reuse the identical shared-portfolio machinery
with a one-symbol universe, so methodology matches the 3-asset headline exactly
(maker entry, 1h sub-bar exits, funding, liquidation, 2bps market-exit
slippage). The portfolio-wide exposure caps in the standard configs are still
applied, so a single-asset multiple here can differ from a truly standalone run.

Run (in-sample 2021-01 → 2025-06):
    PYTHONPATH=. /tmp/tmlvenv/bin/python -m opt.weekly_progression
Output:
    reports/weekly_progression.xlsx

For the out-of-sample forward year (2025-06 → 2026-06) see
``opt.weekly_progression_oos``, which reuses ``run_study``/``build_workbook``.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from opt.multi_portfolio import load_assets
from opt.multi_asset import simulate_multi

START = "2021-01-01"
END = "2025-06-01"
SLIP = 0.0002

# (universe label, [symbols])
UNIVERSES = [
    ("BTC+ETH+SOL", ["BTC", "ETH", "SOL"]),
    ("BTC only", ["BTC"]),
    ("ETH only", ["ETH"]),
]
PROFILES = ["standard", "aggressive"]


def sheet_name(profile: str, universe: str) -> str:
    u = {"BTC+ETH+SOL": "3asset", "BTC only": "BTC", "ETH only": "ETH"}[universe]
    return f"{profile.capitalize()}-{u}"


def run_study(all_assets: dict, symbols: list[str], start: str, end: str):
    """Return (weekly DataFrame, initial_balance, MultiAssetResult)."""
    assets = {s: all_assets[s] for s in symbols}
    result = simulate_multi(
        assets, start, end, slip=SLIP, exit_granularity="sub",
    )
    initial = result.portfolio.initial_balance

    curve = pd.DataFrame(
        {"equity": [p.equity for p in result.equity_curve]},
        index=pd.DatetimeIndex(
            [p.timestamp for p in result.equity_curve]
        ).tz_localize(None),
    )
    # Weekly last mark-to-market equity (week ending Sunday).
    weekly = curve["equity"].resample("W").last().dropna().to_frame()
    weekly["multiple_x"] = weekly["equity"] / initial
    weekly["weekly_return_pct"] = weekly["equity"].pct_change() * 100
    weekly.loc[weekly.index[0], "weekly_return_pct"] = (
        weekly["equity"].iloc[0] / initial - 1
    ) * 100
    running_peak = weekly["equity"].cummax()
    weekly["drawdown_pct"] = (weekly["equity"] - running_peak) / running_peak * 100

    out = weekly.reset_index()
    out.columns = ["Week ending", "Equity ($)", "Multiple (x)",
                   "Weekly return (%)", "Drawdown (%)"]
    out.insert(0, "Week #", range(1, len(out) + 1))
    return out, initial, result


def collect_studies(start: str, end: str, data_end: str | None = None,
                    funding_end: str | None = None) -> list[dict]:
    """Run all profile x universe studies over [start, end]."""
    ctx_kwargs = {}
    if data_end is not None:
        ctx_kwargs["data_end"] = data_end
    if funding_end is not None:
        ctx_kwargs["funding_end"] = funding_end
    studies = []
    for profile in PROFILES:
        all_assets = load_assets("maker", profile, **ctx_kwargs)
        for universe, symbols in UNIVERSES:
            print(f"running {profile:10s} {universe:12s} ...", flush=True)
            weekly, initial, result = run_study(all_assets, symbols, start, end)
            final_x = weekly["Multiple (x)"].iloc[-1]
            studies.append({
                "profile": profile, "universe": universe, "symbols": symbols,
                "weekly": weekly, "initial": initial, "result": result,
                "final_x": final_x,
            })
            print(f"  final {final_x:,.2f}x  weeks {len(weekly)}  "
                  f"reported maxDD {result.max_dd_pct:.2f}%  trades {result.trades}",
                  flush=True)
    return studies


def build_workbook(studies: list[dict], out_path: Path, window_label: str) -> None:
    out_path.parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # --- Summary sheet ---
        summary_rows = []
        for s in studies:
            w = s["weekly"]
            summary_rows.append({
                "Profile": s["profile"].capitalize(),
                "Universe": s["universe"],
                "Window": window_label,
                "Final multiple (x)": s["final_x"],
                "Weeks": len(w),
                "Best week (%)": w["Weekly return (%)"].max(),
                "Worst week (%)": w["Weekly return (%)"].min(),
                "Max drawdown (%)": w["Drawdown (%)"].min(),
                "Reported maxDD (%)": s["result"].max_dd_pct,
                "Trades": s["result"].trades,
                "Win rate (%)": s["result"].win_rate,
            })
        summary = pd.DataFrame(summary_rows)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        # --- Combined weekly multiples (all six side by side) ---
        combined = None
        for s in studies:
            col = f"{s['profile'].capitalize()} {s['universe']}"
            series = s["weekly"].set_index("Week ending")["Multiple (x)"]
            series.name = col
            combined = series.to_frame() if combined is None \
                else combined.join(series, how="outer")
        combined = combined.ffill().reset_index()
        combined.to_excel(writer, sheet_name="Weekly Multiples", index=False)

        # --- Per-study detail sheets ---
        for s in studies:
            s["weekly"].to_excel(
                writer, sheet_name=sheet_name(s["profile"], s["universe"]),
                index=False,
            )

        _format_workbook(writer, studies, summary, combined)


def _format_workbook(writer, studies, summary, combined) -> None:
    from openpyxl.utils import get_column_letter

    money = '#,##0.00'
    mult = '#,##0.00" x"'
    pct = '0.00'

    def autosize(ws, df):
        for j, col in enumerate(df.columns, start=1):
            width = max(len(str(col)),
                        df[col].astype(str).map(len).max() if len(df) else 0)
            ws.column_dimensions[get_column_letter(j)].width = min(max(width + 2, 10), 30)

    ws = writer.sheets["Summary"]
    ws.freeze_panes = "A2"
    autosize(ws, summary)
    for r in range(2, len(summary) + 2):
        ws.cell(r, 4).number_format = mult  # Final multiple
        for c in (6, 7, 8, 9, 11):
            ws.cell(r, c).number_format = pct

    ws = writer.sheets["Weekly Multiples"]
    ws.freeze_panes = "B2"
    autosize(ws, combined)
    for r in range(2, len(combined) + 2):
        for c in range(2, len(combined.columns) + 1):
            ws.cell(r, c).number_format = mult

    for s in studies:
        ws = writer.sheets[sheet_name(s["profile"], s["universe"])]
        df = s["weekly"]
        ws.freeze_panes = "A2"
        autosize(ws, df)
        for r in range(2, len(df) + 2):
            ws.cell(r, 3).number_format = money   # Equity
            ws.cell(r, 4).number_format = mult     # Multiple
            ws.cell(r, 5).number_format = pct      # Weekly return
            ws.cell(r, 6).number_format = pct      # Drawdown


def main() -> None:
    studies = collect_studies(START, END)
    out_path = Path("reports/weekly_progression.xlsx")
    build_workbook(studies, out_path, f"{START} → {END} (in-sample)")
    print(f"\nsaved {out_path}")


if __name__ == "__main__":
    main()
